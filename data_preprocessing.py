#!/usr/bin/env python
# coding: utf-8

# In[2]:


import pandas as pd
from openpyxl import load_workbook

def read_table (file_name, sheet=0, header_row=0):
    split_name = file_name.split(".")
    file_type = split_name[len(split_name)-1]
    if file_type == 'xls':
        file = pd.read_excel (file_name, sheet_name = sheet, header = header_row)
    elif file_type == 'xlsx' :
        wb = load_workbook(file_name)
        ws = wb[wb.sheetnames[sheet]]
        file = pd.DataFrame(ws.values)
        file.columns = file.iloc[header_row]
        file =file[header_row+1:]
    elif file_type == 'csv':
        file = pd.read_csv(file_name)
    else:
        file =  pd.DataFrame(columns=['colum'])
    return file

def merge_str(left_df, right_df, left_col, right_col, how):
    right_df.rename(columns={right_col: left_col}, inplace=True)
    right_df[left_col] = right_df[left_col].astype(str)
    left_df[left_col] = left_df[left_col].astype(str)
    left_df = pd.merge(left_df, right_df, how=how, on=left_col)
    return left_df

def load_trim_attr(file_name, columns, filter_path, column_name, sheet=0, header_row=0, code=0):
    # Read attribute file
    file = read_table(file_name, sheet, header_row)
    # Read filter units file
    filter_units = read_table(filter_path)
    code_header = filter_units.columns[code]
    filter_units[[code_header]]=filter_units[[code_header]].astype(str)
    # Select only code of the geographic units and the number of dwellings as columns
    cols= [list(file.columns)[i] for i in columns]
    file_cols=file[cols]
    # Rename columns
    file_cols.columns = [code_header, cols[1]]
    # Drop NaNs
    file_cols.dropna(inplace =True)
    # Delete blank spaces in the string
    file_cols[code_header] = file_cols[code_header].astype(str)
    file_cols[code_header] = file_cols[code_header].str.replace(' ', '')
    # Apply geographic mask
    result = pd.merge(file_cols, filter_units, how="right", on=code_header)
    # Rename variable of interest
    result.rename(columns={cols[1]: column_name}, inplace=True)
    return result


# In[ ]:




